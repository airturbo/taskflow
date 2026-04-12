from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

out = Path('/Users/turbo/WorkBuddy/20260330162606/docs/visa-principal-acquiring-budget-2026-04-08.xlsx')
wb = Workbook()
wb.remove(wb.active)

thin = Side(style='thin', color='D9E2F3')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill('solid', fgColor='1F4E78')
sub_fill = PatternFill('solid', fgColor='D9EAF7')
input_fill = PatternFill('solid', fgColor='FFF2CC')
summary_fill = PatternFill('solid', fgColor='E2F0D9')
warn_fill = PatternFill('solid', fgColor='FCE4D6')
header_font = Font(color='FFFFFF', bold=True)
sub_font = Font(color='1F1F1F', bold=True)
wrap_top = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
currency_fmt = '$#,##0;($#,##0);-'

rows = [
    ['LIC-01', 'Licensing kick-off', 'Initial Service Fee', 'Licensing', '一次性基线', '是', '是', '是', 'Finance + Biz Owner + Legal/Compliance', 'License 审批启动前', 75000, 75000, 75000, 'Principal 示例值', '是否必须全额预付后才继续审批？', '高', '锁定付款前置条件与审批顺序'],
    ['LIC-02', 'Provisional BID assigned', 'Numeric Licensing Fee', 'Licensing', '一次性基线', '是', '是', '是', 'Implementation Lead + Finance', 'Provisional BID 批下后 / BIN 分配当月', 1000, 1000, 1000, '按示例 2 BIN 估算', '“当月开票”与“BIN go-live 后 billing timing”之间，实际扣款日是哪天？', '中', '确认 invoice date 与 actual collection date'],
    ['IMP-01', 'Implementation kick-off', 'Project Management Fee', 'Implementation', '一次性基线', '是', '是', '是', 'PM/Implementation Lead + Finance', 'Implementation 启动后 6 个月口径', 18000, 18000, 24000, '示例按 6 个月；压力口径按 8 个月', '如果实施超过 6 个月，是否继续按月累计？', '高', '把 implementation 月数纳入预算基线'],
    ['CFG-01', 'CIQ 完成 + config request 提交', 'New Program Installation', 'Configuration', '一次性基线', '是', '是', '是', 'Implementation Lead', 'CIQ 完成并提交配置请求后', 6000, 6000, 9000, '示例按 1 CIB + 1 PCR；压力口径预留额外 numeric', 'Settlement BIN 是否单独收费、是否已包含在当前示例里？', '中', '冻结架构后再提配置'],
    ['CFG-02', 'CIQ 完成 + config request 提交', 'Endpoint Certification Issuance Fee', 'Configuration', '一次性基线', '是', '是', '是', 'Network/Infra Lead + Implementation', 'CIQ 完成并提交配置请求后', 550, 1100, 1100, '建议按 Prod + DR 两张证书预算', 'Production 和 DR 是否各要一张证书？', '中', '确认 endpoint 与证书数量'],
    ['CFG-03', '证书请求', 'eCommerce Digital Certificate Issuance', 'Configuration', '一次性基线', '否', '待定', '否', '3DS Owner + Implementation', '3DS 方案确定后', 550, 550, 1100, '仅自建/持有 3DS Server 时适用', '如果使用第三方 3DS，证书费由谁承担？', '中', '确认 3DS 接入模式与证书 owner'],
    ['TST-01', 'Attended testing 完成', 'Attended Testing Fee', 'Testing', '一次性基线', '是', '是', '是', 'QA/Test Lead + Finance', 'Attended testing 完成后', 12000, 12000, 16000, '示例按 12 小时；压力口径按 16 小时', '超出预估小时数如何加收？是否按实际小时滚动？', '高', '锁定测试时长预估与排期'],
    ['TST-02', '提交 test card request', 'Host Test Cards Package', 'Testing', '一次性基线', '否', '待定', '否', 'QA/Test Lead', '测试准备阶段', 1000, 1000, 2000, '按 1 package 估算', '一个 package 覆盖多少测试场景？需不需要重复买？', '低', '核对测试卡覆盖范围'],
    ['TST-03', '申请 VSTS access', 'Visa Secure Test Suite (VSTS)', 'Testing', '一次性基线', '否', '待定', '否', '3DS Owner + QA', '3DS 测试窗口开启时', 7500, 7500, 15000, '90 天订阅；压力口径按续费一次', '如果 90 天测试未完成，是否续费同价？', '中', '按真实测试窗口申请订阅'],
    ['TST-04', 'IM 确认 proceed with VCX', 'Visa Clearing Exchange License Fee', 'Testing', '一次性基线', '否', '待定', '否', 'Processing/Settlement Lead', 'VCX 方案确认后', 8000, 8000, 8000, '仅需 VCX 时适用', '如果第三方 processor 已覆盖，是否可完全不买？', '中', '比较自建与第三方 processor 成本'],
    ['TST-05', '提交 VTS 工具申请表', 'VisaNet Test System – System License Fee', 'Testing', '一次性基线', '否', '待定', '否', 'QA/Test Lead', 'Host testing 工具申请后', 5000, 5000, 5000, '按单次系统许可证估算', '是一次买断还是本项目一次性授权？', '中', '确认授权性质与复用范围'],
    ['ACC-01', '全部 testing + certification 完成', 'Visa Cloud Connect Access Fee / New EAS Installation Fee', 'Access', '一次性基线', '否', '待定', '否', 'Infra/Network Lead + Finance', 'Endpoint activation 时', 20000, 20000, 20000, '示例按 1 Prod + 1 DR endpoint', '1 Production + 1 DR 是否固定两笔？后续新增 endpoint 如何收费？', '高', '在是否 own endpoint 上做总成本比较'],
    ['EVT-01', '配置改期', 'Configuration Reschedule', 'Event', '可选事件缓冲', '否', '是', '是', 'PM/Implementation Lead', '若改实施日期', 0, 3500, 7000, '建议保留 1 次缓冲；压力口径按 2 次', '变更窗口多晚之前改不收费？', '中', '能不改就不改'],
    ['EVT-02', '配置取消', 'Configuration Cancellation', 'Event', '可选事件缓冲', '否', '是', '否', 'PM/Implementation Lead', '若取消配置请求', 0, 6500, 6500, '默认不纳入；若范围不稳可打开', '哪类取消能豁免？', '高', '范围未冻结前不要提交配置'],
    ['EVT-03', '要求加急配置', 'Configuration Expedite', 'Event', '可选事件缓冲', '否', '是', '否', 'PM/Implementation Lead', '若未满足 lead time 但要求加急', 0, 6500, 13000, '默认不纳入；压力口径按 2 次', '最小 lead time 具体按哪张 Visa 日历？', '高', '把外部依赖前置，避免临时加急'],
    ['EVT-04', '要求加急出证', 'Expedited Endpoint Certification Issuance Fee', 'Event', '可选事件缓冲', '否', '是', '否', 'Infra/Network Lead', '若加急发证', 0, 1000, 2000, '按 1-2 张证书缓冲', '各类证书标准 lead time 各是多少？', '中', '提前锁证书窗口'],
    ['EVT-05', '申请高优先级测试 slot', 'Expedited Testing Requests', 'Event', '可选事件缓冲', '否', '是', '否', 'QA/Test Lead', '若需要测试加急', 0, 3000, 6000, '按 1-2 次缓冲', '2 小时以上如何计费？', '中', '测试排程一次排准'],
    ['EVT-06', '临期取消已排期测试', 'Cancellation Fee for Scheduled Attended Testing', 'Event', '可选事件缓冲', '否', '是', '否', 'QA/Test Lead + PM', '若 24 小时内取消测试', 0, 2000, 4000, '按 1-2 次缓冲', '2 小时 minimum 是否代表少于 2 小时也按 2 小时逻辑收？', '中', '避免临期取消'],
    ['REC-01', '服务开通后', 'Visa Online Access Fee', 'Recurring', '年化经常性', '是', '是', '是', 'Operations + Finance', '服务开通后首年', 5000, 5000, 5000, '按 1 Business ID 年化示例', '从 Provisional BID 还是正式 Business ID 开始计？', '中', '核准起算节点与 Business ID 生效日'],
    ['REC-02', '服务开通后', 'Visa Resolve Online', 'Recurring', '年化经常性', '是', '是', '是', 'Dispute Ops + Finance', 'VROL 开通后首年', 2400, 2400, 2400, '按月费年化', '如果只做 acquirer access，是否仍按完整月费收？', '中', '确认 dispute access 范围'],
    ['REC-03', '年度订阅开通后', 'Testing Self-Service Tools', 'Recurring', '年化经常性', '否', '待定', '否', 'QA/Test Lead', '自测工具开通后首年', 5000, 5000, 5000, '仅使用自测工具时适用', '是否自动续费？', '低', '明确续费和取消机制'],
    ['REC-04', 'VCX 上线后', 'VCX Subscription', 'Recurring', '年化经常性', '否', '待定', '否', 'Processing/Settlement Lead + Finance', 'VCX 方案启用后首年', 6000, 6000, 6000, '按 1 CIB 年化示例', '多个 CIB 如何累计？', '中', '按 CIB 维度估算长期成本'],
    ['REC-05', 'TPA 注册后', 'Third Party Agent Fees', 'Recurring', '年化经常性', '否', '待定', '否', 'Partner Mgmt + Compliance + Finance', 'TPA 启用后首年', 2500, 5000, 10000, '示例取 5,000；按 merchant volume 分档', '最终 volume 档位以哪版 Fee Schedule 为准？', '中', '补拉最新 Fee Schedule 档位表'],
    ['REC-06', '直接清算关系成立后', 'FTSRE Fee', 'Recurring', '年化经常性', '否', '待定', '否', 'Treasury/Settlement Ops', '直接清算关系建立后首年', 0, 0, 4200, '单币种国际结算可不需要；压力口径按 1 点位年化', '单币种国际结算可否完全免掉？', '中', '确认 settlement currency 架构'],
    ['REC-07', '自建 endpoint 上线后', 'Visa Cloud Connect Monthly Maintenance Fee', 'Recurring', '年化经常性', '否', '待定', '否', 'Infra/Network Lead + Finance', '自建 endpoint 投产后首年', 60000, 60000, 600000, '文档月费 5,000-50,000；建议先用下限建模，压力口径按上限', 'volume 阶梯怎么切档？', '高', '尽快拿到 volume pricing 规则'],
    ['POST-01', '宽限期结束', 'Quarterly Minimum Fee', 'Recurring', '年化经常性', '是', '是', '是', 'Finance + Biz Owner', 'P-BID 后 3 个季度宽限结束后首年', 30000, 30000, 30000, '按 7,500/季年化', '若 volume 提前启动，具体从哪个账期开始收？', '高', '记录 P-BID 生效日并做季度倒排'],
    ['POST-02', '评级不足的运营期', 'Risk Monitoring Fee', 'Recurring', '年化经常性', '否', '待定', '否', 'Treasury/Risk + Finance', '无合格评级时首年', 16000, 16000, 16000, '4,000/季年化示例', '如果后续拿到评级，是否次季度起停收？', '高', '确认评级豁免与停收规则'],
    ['POST-03', '开始跑交易后', 'License / Service / Network / Transaction Fees', 'Transaction', '待补费率', '是', '是', '否', 'Finance + Payments Ops', '交易开始后', 0, 0, 0, '本 PDF 未展开，需另拉 Fee Schedule', '最新 Fee Schedule 明细能否单独拉表？', '高', '单独建立长期利润测算底表'],
]

yes_no = '是,否,待定'
risk_opts = '高,中,低'


def add_header(ws, row_idx, titles):
    for col_idx, title in enumerate(titles, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border


def style_sheet(ws, widths, freeze='A2'):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = True
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_top
            cell.border = border

ws = wb.create_sheet('预算说明')
ws['A1'] = 'Visa Principal Acquiring 预算版 Excel'
ws['A1'].font = Font(size=14, bold=True)
ws['A1'].fill = summary_fill
info = [
    ['来源文件', 'Projected Setup Cost  Principal Acquiring070426.pdf'],
    ['预算口径', '一次性基线 + 可选事件缓冲 + 年化经常性；交易级费率单独待补'],
    ['建议用法', '先在“预算清单”确定适用项和是否纳入预算，再看“预算汇总”读当前预算与压力预算'],
    ['关键字段', '“是否适用”“是否纳入预算”可直接改；当前纳入预算和压力预算会自动汇总'],
    ['注意', '带范围的项目默认采用建议预算和压力预算双口径；最新收费仍以 Visa Fee Schedule / Visa Online / MIDAS 正式回复为准'],
]
add_header(ws, 3, ['项目', '内容'])
for r, row in enumerate(info, 4):
    for c, value in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=value)
        if c == 1:
            cell.fill = sub_fill
            cell.font = sub_font
style_sheet(ws, {1: 18, 2: 104}, freeze='A3')

ws = wb.create_sheet('预算清单')
headers = ['ID', '阶段/节点', '费用项', '费用分组', '预算口径', '是否必选', '是否适用', '是否纳入预算', '预算 Owner', '预计发生时点', '金额下限(USD)', '建议预算(USD)', '压力预算(USD)', '当前纳入预算(USD)', '当前压力预算(USD)', '备注/假设', '待确认问题', '风险等级', '下一步动作']
add_header(ws, 1, headers)
for r, row in enumerate(rows, 2):
    for c, value in enumerate(row, 1):
        ws.cell(row=r, column=c, value=value)
    ws.cell(row=r, column=14, value=f'=IF(OR(G{r}="否",H{r}="否"),0,L{r})')
    ws.cell(row=r, column=15, value=f'=IF(OR(G{r}="否",H{r}="否"),0,M{r})')
    ws.cell(row=r, column=7).fill = input_fill
    ws.cell(row=r, column=8).fill = input_fill
    ws.cell(row=r, column=18).fill = warn_fill if row[16] == '高' else (input_fill if row[16] == '中' else summary_fill)
for col in ['K', 'L', 'M', 'N', 'O']:
    for row_idx in range(2, len(rows) + 2):
        ws[f'{col}{row_idx}'].number_format = currency_fmt
ws.auto_filter.ref = f'A1:S{len(rows)+1}'
style_sheet(ws, {1: 10, 2: 24, 3: 34, 4: 16, 5: 16, 6: 10, 7: 10, 8: 12, 9: 28, 10: 26, 11: 14, 12: 14, 13: 14, 14: 16, 15: 16, 16: 28, 17: 42, 18: 10, 19: 30})
yn_dv = DataValidation(type='list', formula1=f'"{yes_no}"', allow_blank=True)
risk_dv = DataValidation(type='list', formula1=f'"{risk_opts}"', allow_blank=True)
ws.add_data_validation(yn_dv)
ws.add_data_validation(risk_dv)
yn_dv.add(f'F2:H{len(rows)+1}')
risk_dv.add(f'R2:R{len(rows)+1}')

ws = wb.create_sheet('预算汇总')
add_header(ws, 1, ['指标', '当前预算(USD)', '压力预算(USD)', '说明'])
summary_rows = [
    ['一次性基线', '=SUMIFS(预算清单!$N:$N,预算清单!$E:$E,"一次性基线")', '=SUMIFS(预算清单!$O:$O,预算清单!$E:$E,"一次性基线")', 'License / Implementation / Testing / Access 等项目型成本'],
    ['可选事件缓冲', '=SUMIFS(预算清单!$N:$N,预算清单!$E:$E,"可选事件缓冲")', '=SUMIFS(预算清单!$O:$O,预算清单!$E:$E,"可选事件缓冲")', '用于改期、取消、加急等缓冲'],
    ['年化经常性', '=SUMIFS(预算清单!$N:$N,预算清单!$E:$E,"年化经常性")', '=SUMIFS(预算清单!$O:$O,预算清单!$E:$E,"年化经常性")', '按首年 run-rate 年化'],
    ['待补费率', '=SUMIFS(预算清单!$N:$N,预算清单!$E:$E,"待补费率")', '=SUMIFS(预算清单!$O:$O,预算清单!$E:$E,"待补费率")', '交易级费率暂未纳入'],
    ['总预算（不含待补费率）', '=SUM(B2:B4)', '=SUM(C2:C4)', '当前最可执行预算口径'],
    ['高风险已纳入预算', '=SUMIFS(预算清单!$N:$N,预算清单!$R:$R,"高")', '=SUMIFS(预算清单!$O:$O,预算清单!$R:$R,"高")', '优先盯防高风险项'],
    ['未纳入预算的建议预算', '=SUMIFS(预算清单!$L:$L,预算清单!$H:$H,"否")', '=SUMIFS(预算清单!$M:$M,预算清单!$H:$H,"否")', '这些是潜在漏预算区'],
]
for r, row in enumerate(summary_rows, 2):
    for c, value in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=value)
        if c == 1:
            cell.fill = sub_fill
            cell.font = sub_font
for col in ['B', 'C']:
    for row_idx in range(2, len(summary_rows) + 2):
        ws[f'{col}{row_idx}'].number_format = currency_fmt
ws['A10'] = '建议读法'
ws['A10'].fill = summary_fill
ws['A10'].font = sub_font
ws['B10'] = '先看“总预算（不含待补费率）”把项目预算框住，再把“未纳入预算的建议预算”压小，最后单独补 transaction fee schedule。'
style_sheet(ws, {1: 22, 2: 18, 3: 18, 4: 56}, freeze='A2')

ws = wb.create_sheet('预算时间线')
add_header(ws, 1, ['时间点', '节点', '预算重点', '对应费用', '管理动作'])
time_rows = [
    ['T0', 'Licensing review kick-off', '先锁最前置现金流', 'Initial Service Fee', '立项时直接入预算'],
    ['T1', 'Provisional BID 批下', '记录宽限期时钟起点', 'Numeric Licensing Fee；Quarterly Minimum Fee 时钟', '确认 P-BID 生效日'],
    ['T2', 'Implementation kick-off', '防止实施拖长吃预算', 'Project Management Fee', '把月数写进里程碑'],
    ['T3', 'CIQ / 配置请求提交', '配置型一次性成本集中出现', 'Program Installation；各类证书费', '范围冻结后再提交'],
    ['T4', '测试准备与测试执行', '测试工具和人力成本混合爆发', 'VSTS；VTS；VCX；Attended Testing', '一次排准测试，减少返工'],
    ['T5', 'Endpoint activation', '自建 endpoint 大额接入费落地', 'Cloud Connect Access / EAS Installation', '与第三方方案做总成本对比'],
    ['T6', 'BIN go-live', '集中看前期开票与自动扣收', '多项 setup fee billing timing', '做上线前 billing checklist'],
    ['T7', 'Go-live 后', '从项目支出切到运营支出', 'Visa Online；VROL；VCX；Maintenance', '把 recurring fee 纳入 run-rate'],
    ['T8', '宽限期结束', '底座成本正式启动', 'Quarterly Minimum Fee', '提前倒排季度账期'],
    ['T9', '长期运营', '看利润率而不是只看 setup fee', 'Recurring fee + transaction-level fees', '持续做 run-rate 与 margin 复盘'],
]
for r, row in enumerate(time_rows, 2):
    for c, value in enumerate(row, 1):
        ws.cell(row=r, column=c, value=value)
style_sheet(ws, {1: 10, 2: 28, 3: 34, 4: 40, 5: 36})
ws.auto_filter.ref = f'A1:E{len(time_rows)+1}'

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

wb.save(out)
print(out)
