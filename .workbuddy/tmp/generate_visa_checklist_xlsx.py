from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

out = Path('/Users/turbo/WorkBuddy/20260330162606/.workbuddy/tmp/visa-principal-acquiring-checklist-2026-04-08.xlsx')
wb = Workbook()
wb.remove(wb.active)

thin = Side(style='thin', color='D9E2F3')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill('solid', fgColor='1F4E78')
sub_fill = PatternFill('solid', fgColor='D9EAF7')
check_fill = PatternFill('solid', fgColor='FFF2CC')
summary_fill = PatternFill('solid', fgColor='E2F0D9')
warn_fill = PatternFill('solid', fgColor='FCE4D6')
header_font = Font(color='FFFFFF', bold=True)
sub_font = Font(color='1F1F1F', bold=True)
wrap_top = Alignment(wrap_text=True, vertical='top')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

fee_rows = [
    ['LIC-01', 'Licensing kick-off', 'Initial Service Fee', '前置牌照费', '一次性；Principal 示例 75,000 美元', 'Licensing review kick-off 即触发并开票', '需尽早 wire transfer 给 Visa，以推进 license approval', 'Finance + Biz Owner + Legal/Compliance', 'License application 正式启动', '高', '未开始', '是不是必须全额预付后才继续审批？是否允许分批？', '确认付款前置条件并锁定审批节奏', '文档最前置费用'],
    ['LIC-02', 'Provisional BID assigned', 'Numeric Licensing Fee', '牌照/标识费', '500 / BIN；示例 2 BIN = 1,000 美元', 'Provisional BID 分配后，MIDAS 才能继续 assign BIN；当月开票', '文档写 After BIN go live', 'Implementation Lead + Finance', 'Provisional BID 批下', '中', '未开始', '“当月开票”与“BIN go-live 后 billing timing”之间，实际扣款日是哪天？', '确认 invoice date 与 actual collection date', '与 BIN 分配直接相关'],
    ['IMP-01', 'Implementation kick-off', 'Project Management Fee', '实施月费', '3,000 / 项目 / 月；示例 6 个月 = 18,000 美元', 'Implementation project kick-off，IM 入场后开始按月计', '文档写 After BIN go live', 'PM/Implementation Lead + Finance', 'Implementation manager 已指派', '高', '未开始', '如果实施超过 6 个月，是否继续按月累计？', '把 implementation 时长纳入预算基线', '拖期会直接放大成本'],
    ['CFG-01', 'CIQ 完成 + config request 提交', 'New Program Installation', '一次性配置费', 'PCR 3,000；CIB 3,000；示例 6,000 美元', 'CIQ 完成且 IM 提交 configuration request 后触发', 'After BIN go live', 'Implementation Lead', 'CIQ 完成、numeric 已就位', '中', '未开始', 'Settlement BIN 是否单独收费、是否已包含在当前示例里？', '冻结架构范围后再提配置', '多一个 PCR/CIB 成本就多一层'],
    ['CFG-02', 'CIQ 完成 + config request 提交', 'Endpoint Certification Issuance Fee', '证书费', '550 / 证书', 'CIQ 完成并提交配置请求后触发', 'After BIN go live', 'Network/Infra Lead + Implementation', '需加密连接 VisaNet', '中', '未开始', 'Production 和 DR 是否各要一张证书？', '明确 endpoint 与证书数量', '与 VisaNet 加密接入相关'],
    ['CFG-03', '证书请求', 'eCommerce Digital Certificate Issuance', '证书费', '550 / 证书', '请求 3DS Server certificate 时触发', '文档整体归在 After BIN go live 逻辑', '3DS Owner + Implementation', '自建或接管 3DS Server', '中', '未开始', '如果使用第三方 3DS，证书费由谁承担？', '确认 3DS 接入模式与证书 owner', '适用于 3DS Server connectivity'],
    ['TST-01', 'Attended testing 完成', 'Attended Testing Fee', '测试服务费', '1,000 / 小时；示例 12,000 美元', 'attended testing 完成并统计总小时数后开票', 'After BIN go live', 'QA/Test Lead + Finance', 'Visa 测试团队参与', '高', '未开始', '超出预估小时数如何加收？是否按实际小时滚动？', '锁定测试时长预估与排期', '实际小时不确定性高'],
    ['TST-02', '提交 test card request', 'Host Test Cards Package', '测试物料费', '1,000 / package', '客户提交 acquiring testing test cards request 后触发', 'After BIN go live', 'QA/Test Lead', '需要实体/测试卡资源', '低', '未开始', '一个 package 覆盖多少测试场景？需不需要重复买？', '核对测试卡覆盖范围', '物料型成本'],
    ['TST-03', '申请 VSTS access', 'Visa Secure Test Suite (VSTS)', '测试工具订阅', '7,500 / 90 天订阅', '请求 VSTS access 即触发', '文档归在 After BIN go live', '3DS Owner + QA', '自有 3DS Server', '中', '未开始', '如果 90 天测试未完成，是否续费同价？', '按真实测试窗口申请订阅', '仅自有 3DS Server 适用'],
    ['TST-04', 'IM 确认 proceed with VCX', 'Visa Clearing Exchange License Fee', '测试/接入许可证', '一次性 8,000', 'IM 确认启用 VCX 时触发', 'After BIN go live', 'Processing/Settlement Lead', '需要 VCX', '中', '未开始', '如果第三方 processor 已覆盖，是否可完全不买？', '比较自建与第三方 processor 成本', 'BASE II VCX software'],
    ['TST-05', '提交 VTS 工具申请表', 'VisaNet Test System – System License Fee', '测试工具许可证', '一次性 5,000', 'VTS request form 提交时触发', 'After BIN go live', 'QA/Test Lead', '需用 Visa 提供 host testing 工具', '中', '未开始', '是一次买断还是本项目一次性授权？', '确认授权性质与复用范围', '用于 host testing'],
    ['ACC-01', '全部 testing + certification 完成', 'Visa Cloud Connect Access Fee / New EAS Installation Fee', '接入开通费', '10,000 / endpoint；示例 20,000', 'endpoint activation 时触发', 'activation 基本就是正式收取节点', 'Infra/Network Lead + Finance', '自建 endpoint、完成认证与测试', '高', '未开始', '1 Production + 1 DR 是否固定两笔？后续新增 endpoint 如何收费？', '在是否 own endpoint 上做总成本比较', '自建 endpoint 模式关键大项'],
    ['EVT-01', '配置已被 Visa acknowledge 后要求改期', 'Configuration Reschedule', '事件费', '3,500 / 次', '改实施日期时触发', '事件发生即收', 'PM/Implementation Lead', '已发 acknowledgement', '中', '未开始', '变更窗口多晚之前改不收费？', '上线窗口冻结前不要轻易提交', '典型可避免成本'],
    ['EVT-02', '配置已被 Visa acknowledge 后取消', 'Configuration Cancellation', '事件费', '6,500 / 次', '取消配置请求时触发', '事件发生即收', 'PM/Implementation Lead', '已发 acknowledgement', '高', '未开始', '哪类取消能豁免？', '避免在范围未冻结前提交配置', '高惩罚性费用'],
    ['EVT-03', '不满足标准 lead time 但要求加急', 'Configuration Expedite', '事件费', '6,500 / 次', '提出加急时触发', '事件发生即收', 'PM/Implementation Lead', '紧急配置诉求', '高', '未开始', '最小 lead time 具体按哪张 Visa 日历？', '把外部依赖前置，避免临时加急', '项目纪律成本'],
    ['EVT-04', '要求加急出证', 'Expedited Endpoint Certification Issuance Fee', '事件费', '1,000 / 证书', '提出加急发证时触发', '事件发生即收', 'Infra/Network Lead', '证书 lead time 不满足', '中', '未开始', '各类证书标准 lead time 各是多少？', '提前锁证书窗口', '证书类加急成本'],
    ['EVT-05', '申请高优先级测试 slot', 'Expedited Testing Requests', '事件费', '3,000 / 次', '申请加急测试时触发', '事件发生即收', 'QA/Test Lead', '需要更高优先级 slot', '中', '未开始', '2 小时以上如何计费？', '测试排程一次排准', '测试资源加急成本'],
    ['EVT-06', '已排期测试在 24 小时内取消', 'Cancellation Fee for Scheduled Attended Testing', '事件费', '2,000 / 次', '<24h 取消时触发', '事件发生即收', 'QA/Test Lead + PM', '测试 slot 已锁定', '中', '未开始', '2 小时 minimum 是否代表少于 2 小时也按 2 小时逻辑收？', '避免临期取消已排期测试', '测试取消惩罚'],
    ['REC-01', '服务开通后', 'Visa Online Access Fee', '运营 recurring fee', '1,250 / 季 / Business ID；示例年化 5,000', 'Visa Online access / Business ID 有效后开始', '按季持续', 'Operations + Finance', 'Visa Online 已开通', '中', '未开始', '从 Provisional BID 还是正式 Business ID 开始计？', '核准起算节点与 Business ID 生效日', 'B2B 门户 access fee'],
    ['REC-02', '服务开通后', 'Visa Resolve Online', '运营 recurring fee', '200 / 月；示例年化 2,400', 'VROL access 开通后', '按月持续', 'Dispute Ops + Finance', '争议处理链路启用', '中', '未开始', '如果只做 acquirer access，是否仍按完整月费收？', '确认 dispute access 范围', '争议管理基础包'],
    ['REC-03', '年度订阅开通后', 'Testing Self-Service Tools', '运营 recurring fee', '5,000 / 年', '自测工具订阅开通后', '按年持续', 'QA/Test Lead', '使用自测工具', '低', '未开始', '是否自动续费？', '明确续费和取消机制', '年度工具费'],
    ['REC-04', 'VCX 上线后', 'VCX Subscription', '运营 recurring fee', '500 / 月 / CIB；示例年化 6,000', 'VCX subscription 生效后', '按月持续', 'Processing/Settlement Lead + Finance', '使用 VCX', '中', '未开始', '按 CIB 收费，多个 CIB 如何累计？', '按 CIB 维度估算长期成本', 'VCX 运维订阅'],
    ['REC-05', 'TPA 注册后', 'Third Party Agent Fees', '运营 recurring fee', '年费；按 merchant volume 分档', 'TPA 注册且开始提供支付相关服务后', '按年持续', 'Partner Mgmt + Compliance + Finance', '存在第三方服务商', '中', '未开始', '最终 volume 档位以哪版 Fee Schedule 为准？', '补拉最新 Fee Schedule 档位表', 'OCR 档位有轻微错位'],
    ['REC-06', '直接清算关系成立后', 'FTSRE Fee', '运营 recurring fee', '350 / 月 / 额外 funds transfer point', 'direct settlement relationship + 额外 transfer point 出现后', '按月持续', 'Treasury/Settlement Ops', '多资金划转点', '中', '未开始', '单币种国际结算可否完全免掉？', '确认 settlement currency 架构', '单币种国际结算通常可不需要'],
    ['REC-07', '自建 endpoint 上线后', 'Visa Cloud Connect Monthly Maintenance Fee', '运营 recurring fee', '5,000–50,000 / 月', '自建 endpoint 投产后', '按月持续', 'Infra/Network Lead + Finance', 'own endpoint + volume', '高', '未开始', 'volume 阶梯怎么切档？', '尽快拿到 volume pricing 规则', '长期 burden 可能很重'],
    ['POST-01', '宽限期结束', 'Quarterly Minimum Fee', '运营底座费', '7,500 / 季；示例年化 30,000', 'Provisional Business ID 后 3 个季度 grace period 结束；若 volume 更早启动则可能更早开始', '按季持续', 'Finance + Biz Owner', 'Principal 资格仍在、宽限结束', '高', '未开始', '若 volume 提前启动，具体从哪个账期开始收？', '记录 P-BID 生效日并做季度倒排', '即使没量也可能开始收'],
    ['POST-02', '运营期但信用评级不足', 'Risk Monitoring Fee', '运营风控费', '4,000 / 季；示例年化 16,000', 'post-implementation 运营期内仍不满足国际/全球评级条件时', '按季持续', 'Treasury/Risk + Finance', '无 S&P/Moody’s/Fitch 合格评级', '高', '未开始', '如果后续拿到评级，是否次季度起停收？', '确认评级豁免与停收规则', '对 non-FI 尤其敏感'],
    ['POST-03', '开始跑交易后', 'License / Service / Network / Transaction Fees', '交易级费率', '按 Fee Schedule', '一旦交易和服务真实发生即开始', '随交易 / 服务使用持续结算', 'Finance + Payments Ops', 'Go-live、交易发生', '高', '未开始', '最新 Fee Schedule 明细能否单独拉表？', '必须拉一版长期利润测算底表', '本 PDF 未展开明细'],
]

timeline_rows = [
    ['T0', 'License 立项 / Licensing review kick-off', 'Initial Service Fee', '最前置、最刚性，像入场门票；可能直接卡 license approval', '立项前就纳入预算，别等技术快准备完才想起这笔', '前置现金流'],
    ['T1', 'Provisional BID 批下', 'Numeric Licensing Fee；Quarterly Minimum Fee 宽限期开始计时', '进入“实施准备已具备资格”的状态，后续大量收费开始有资格发生', '记录 Provisional BID/Business ID 的准确生效日期', '宽限期时钟启动'],
    ['T2', 'Implementation kick-off', 'Project Management Fee（月度）', '典型时间成本开始流出；项目越拖，成本越高', '把 implementation 月数纳入预算与里程碑管理', '拖期敏感'],
    ['T3', 'CIQ 完成、配置请求正式提交', 'New Program Installation；Endpoint Certification；eCommerce Digital Certificate', '配置型一次性费用开始集中出现', '在 CIQ 提交前冻结目标架构，避免后续改期/取消/加急', '配置密集期'],
    ['T4', '测试准备与测试执行', 'Host Test Cards；VSTS；VTS；VCX License；Attended Testing；可能还有加急/取消事件费', '工具费 + 人工测试费混合爆发，其中 attended testing 存在小时数不确定性', '先确认哪些工具真的要买，并把测试排程一次排准', '波动区间较大'],
    ['T5', 'Endpoint activation / 接入真正开通', 'Visa Cloud Connect Access Fee / New EAS Installation Fee', '自建 endpoint 模式下最明确的大额一次性接入费', '若不是必须 own endpoint，要跟第三方 processor 方案做总成本比较', '架构选择关键点'],
    ['T6', 'BIN go-live', '前面积累的 setup fee 进入正式收/结；Auto-collection of fees；Post go-live 2 周监控开始', '费用开始在 go-live 前后汇聚', '上线前做 billing checklist：已开票未扣 / 会自动收 / 首月首季 recurring fee', '结算汇聚点'],
    ['T7', 'Go-live 后 0–2 周', 'Visa Online；Visa Resolve Online；可能还有自测工具 / VCX 月费 / Cloud Connect maintenance', '从一次性 setup 支出切换到持续运营型支出', '财务模型从项目成本切到运营成本', 'OPEX 启动'],
    ['T8', 'Provisional Business ID 后第 3 个季度结束', 'Quarterly Minimum Fee（若未因 volume 提前触发）', '运营期底座成本开始启动，即使没量也可能开始收', '不能按“有量才付”理解，需前置纳入预算', '底座成本启动'],
    ['T9', '长期运营阶段', 'Quarterly Minimum Fee；Risk Monitoring Fee；Visa Online；VROL；VCX；TPA；FTSRE；Cloud Connect Maintenance；Transaction-level fees', '进入“固定 recurring fee + 可变交易费”双层模型', '长期盯固定底费、架构选择成本、交易型成本三大类', '利润率核心区'],
]

questions = [
    [1, 'Initial Service Fee 是否必须全额到账后，license approval 才继续？', 'Initial Service Fee', 'Finance + Legal/Compliance', '高', '未开始', '向 Visa / MIDAS 明确付款前置条件'],
    [2, 'Numeric Licensing Fee 是“当月开票、go-live 后扣款”还是“go-live 后才真正出账”？', 'Numeric Licensing Fee', 'Finance + Implementation', '高', '未开始', '核对 invoice date、billing timing、collection date'],
    [3, 'Project Management Fee 超过 6 个月是否继续线性按月累计？', 'Project Management Fee', 'PM + Finance', '高', '未开始', '确认超过示例周期后的计费规则'],
    [4, 'Production 与 DR endpoint / certificate 是否分别收费、分别维护？', 'Endpoint / Certificate 相关费用', 'Infra/Network + Implementation', '中', '未开始', '拉一张 endpoint/certificate 数量清单'],
    [5, 'Quarterly Minimum Fee 若 volume 提前启动，具体从哪个账期开始收？', 'Quarterly Minimum Fee', 'Finance + Biz Owner', '高', '未开始', '用 Provisional Business ID 生效日反推账期'],
    [6, 'Risk Monitoring Fee 若后续补齐评级，是否从下一季度停止？', 'Risk Monitoring Fee', 'Treasury/Risk + Finance', '高', '未开始', '确认评级豁免与停收的账期规则'],
    [7, 'Cloud Connect Monthly Maintenance Fee 的 volume 阶梯和切档规则是什么？', 'Visa Cloud Connect Monthly Maintenance Fee', 'Infra/Network + Finance', '高', '未开始', '向 Visa 索取 volume pricing 明细'],
    [8, '交易型收费（License/Service/Network/International Fees）最新 Fee Schedule 能否单独拉表？', 'Transaction-level fees', 'Finance + Payments Ops', '高', '未开始', '单独建立长期利润测算底表'],
]

legend_rows = [
    ['文件用途', '给 implementation / finance / business / compliance 做内部对齐和逐项确认'],
    ['使用方式', '先看“收费时间线”理解节奏，再在“费用清单”逐项补内部确认状态和备注，最后用“待确认问题”推进对外确认'],
    ['状态建议值', '未开始 / 进行中 / 已确认 / 待 Visa 回复 / 不适用'],
    ['风险等级', '高 = 现金流影响大或规则不清；中 = 需要确认但可控；低 = 相对小额或边缘场景'],
    ['重点盯防', '前置 licensing 付款、implementation 拖期、事件费、Quarterly Minimum Fee、Cloud Connect maintenance、交易级费率'],
    ['注意', '本表基于 PDF 摘要整理，最新收费仍以 Visa Fee Schedule / Visa Online / MIDAS 明确回复为准'],
]

status_options = '未开始,进行中,已确认,待 Visa 回复,不适用'
priority_options = '高,中,低'


def style_sheet(ws, widths, freeze='A2'):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = True
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_top
            cell.border = border


def add_header(ws, row_idx, titles):
    for col_idx, title in enumerate(titles, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

ws = wb.create_sheet('使用说明')
ws['A1'] = 'Visa Principal Acquiring Excel 风格清单'
ws['A1'].font = Font(size=14, bold=True, color='1F1F1F')
ws['A1'].fill = summary_fill
ws['A2'] = '来源文件'
ws['B2'] = 'Projected Setup Cost  Principal Acquiring070426.pdf'
ws['A3'] = '整理日期'
ws['B3'] = '2026-04-08'
ws['A4'] = '说明'
ws['B4'] = '这是一份可直接内部协作使用的 Excel 工作底表，重点帮助团队理解费用触发节点、结算逻辑、责任分工和待确认事项。'
start = 6
add_header(ws, start, ['项目', '内容'])
for r, row in enumerate(legend_rows, start + 1):
    for c, value in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=value)
        cell.fill = sub_fill if c == 1 else PatternFill(fill_type=None)
style_sheet(ws, {1: 20, 2: 100}, freeze='A6')

ws = wb.create_sheet('费用清单')
headers = ['ID', '阶段/节点', '费用项', '费用类型', '金额/频率', '开始计算/触发节点', '收取/结算节点', '建议责任人', '关键依赖', '风险等级', '内部确认状态', '待确认问题', '下一步动作', '备注']
add_header(ws, 1, headers)
for r, row in enumerate(fee_rows, 2):
    for c, value in enumerate(row, 1):
        ws.cell(row=r, column=c, value=value)
    ws.cell(row=r, column=10).fill = warn_fill if row[9] == '高' else (check_fill if row[9] == '中' else summary_fill)
    ws.cell(row=r, column=11).fill = check_fill
ws.auto_filter.ref = f'A1:N{len(fee_rows)+1}'
style_sheet(ws, {1: 11, 2: 26, 3: 34, 4: 18, 5: 28, 6: 34, 7: 30, 8: 30, 9: 26, 10: 10, 11: 14, 12: 42, 13: 34, 14: 22})
fee_status_dv = DataValidation(type='list', formula1=f'"{status_options}"', allow_blank=True)
priority_dv = DataValidation(type='list', formula1=f'"{priority_options}"', allow_blank=True)
ws.add_data_validation(fee_status_dv)
ws.add_data_validation(priority_dv)
fee_status_dv.add(f'K2:K{len(fee_rows)+1}')
priority_dv.add(f'J2:J{len(fee_rows)+1}')

ws = wb.create_sheet('收费时间线')
headers = ['时间点', '节点', '会开始花的钱/触发费用', '现金流特征', '管理重点', '标签']
add_header(ws, 1, headers)
for r, row in enumerate(timeline_rows, 2):
    for c, value in enumerate(row, 1):
        ws.cell(row=r, column=c, value=value)
    ws.cell(row=r, column=6).fill = summary_fill
ws.auto_filter.ref = f'A1:F{len(timeline_rows)+1}'
style_sheet(ws, {1: 10, 2: 30, 3: 48, 4: 40, 5: 42, 6: 16})

ws = wb.create_sheet('待确认问题')
headers = ['序号', '待确认问题', '关联费用/节点', '建议责任人', '优先级', '推进状态', '下一步动作']
add_header(ws, 1, headers)
for r, row in enumerate(questions, 2):
    for c, value in enumerate(row, 1):
        ws.cell(row=r, column=c, value=value)
    ws.cell(row=r, column=5).fill = warn_fill
    ws.cell(row=r, column=6).fill = check_fill
ws.auto_filter.ref = f'A1:G{len(questions)+1}'
style_sheet(ws, {1: 8, 2: 46, 3: 28, 4: 28, 5: 10, 6: 14, 7: 34})
q_priority_dv = DataValidation(type='list', formula1=f'"{priority_options}"', allow_blank=True)
q_status_dv = DataValidation(type='list', formula1=f'"{status_options}"', allow_blank=True)
ws.add_data_validation(q_priority_dv)
ws.add_data_validation(q_status_dv)
q_priority_dv.add(f'E2:E{len(questions)+1}')
q_status_dv.add(f'F2:F{len(questions)+1}')

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1 or (ws.title == '使用说明' and cell.row == 6):
                continue
            if ws.title == '使用说明' and cell.column == 1 and cell.row >= 7:
                cell.font = sub_font
            if ws.title != '使用说明' and cell.column in {1, 10, 11}:
                cell.alignment = center if cell.column in {1, 10, 11} else wrap_top

wb.save(out)
print(out)
